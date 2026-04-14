from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, JSONResponse
import json, csv, io, zipfile
from typing import Optional

from app.core.database import get_db
from app.core.security import get_current_user

router = APIRouter(prefix="/export")


async def get_noeuds_telecom(db, commune=None, statut=None):
    q = "SELECT id::text, nom_unique, type_noeud, etat, commune, ST_Y(geom) AS lat, ST_X(geom) AS lng FROM noeud_telecom WHERE geom IS NOT NULL"
    params = []
    if statut: q += f" AND etat = ${len(params)+1}"; params.append(statut)
    if commune: q += f" AND commune ILIKE ${len(params)+1}"; params.append(f"%{commune}%")
    return await db.fetch(q, *params)

async def get_liens_telecom(db, statut=None):
    q = """SELECT lt.id::text, lt.nom_unique, lt.type_lien, lt.etat, lt.nb_fibres, lt.longueur_m,
                  nd.nom_unique AS noeud_dep, na.nom_unique AS noeud_arr,
                  COALESCE(ST_AsGeoJSON(lt.geom), ST_AsGeoJSON(ST_MakeLine(nd.geom, na.geom)))::json AS geometry
           FROM lien_telecom lt
           JOIN noeud_telecom nd ON nd.id = lt.id_noeud_depart
           JOIN noeud_telecom na ON na.id = lt.id_noeud_arrivee"""
    params = []
    if statut: q += f" WHERE lt.etat = ${len(params)+1}"; params.append(statut)
    return await db.fetch(q, *params)

async def get_logements(db, commune=None, statut=None):
    q = "SELECT id::text, nom_unique, adresse, commune, type_logement, statut_ftth, nb_el_reel, nb_el_raccordables, nb_el_raccordes, ST_Y(geom) AS lat, ST_X(geom) AS lng FROM logement WHERE geom IS NOT NULL"
    params = []
    if statut: q += f" AND statut_ftth = ${len(params)+1}"; params.append(statut)
    if commune: q += f" AND commune ILIKE ${len(params)+1}"; params.append(f"%{commune}%")
    return await db.fetch(q, *params)

async def get_zones(db):
    return await db.fetch("SELECT id::text, nom, code, type_zone, statut, nb_clients_actifs, ST_AsGeoJSON(geom)::json AS geometry FROM zone_influence WHERE geom IS NOT NULL")


def rows_to_geojson(rows, geom_key="geometry", lat_key="lat", lng_key="lng"):
    features = []
    for r in rows:
        d = dict(r)
        if geom_key in d and d[geom_key]:
            geom = d.pop(geom_key)
        elif lat_key in d and lng_key in d and d.get(lat_key) and d.get(lng_key):
            geom = {"type": "Point", "coordinates": [d.pop(lng_key), d.pop(lat_key)]}
            d.pop(lat_key, None)
        else:
            continue
        # Convert non-serializable types
        props = {k: (str(v) if not isinstance(v, (str, int, float, bool, type(None))) else v)
                 for k, v in d.items()}
        features.append({"type": "Feature", "geometry": geom, "properties": props})
    return {"type": "FeatureCollection", "features": features}


def rows_to_csv(rows, headers=None):
    if not rows: return ""
    buf = io.StringIO()
    h = headers or list(dict(rows[0]).keys())
    w = csv.DictWriter(buf, fieldnames=h, extrasaction='ignore')
    w.writeheader()
    for r in rows:
        row = {k: str(v) if v is not None else '' for k, v in dict(r).items()}
        w.writerow(row)
    return buf.getvalue()


@router.get("")
async def exporter(
    format: str = Query("geojson", description="geojson|csv|xlsx|shapefile|kml"),
    couches: str = Query("noeuds_telecom", description="Couches séparées par virgules"),
    commune: Optional[str] = None,
    statut: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    db=Depends(get_db)
):
    couche_list = [c.strip() for c in couches.split(",")]

    # Collecter les données
    all_data = {}
    for couche in couche_list:
        try:
            if couche == "noeuds_telecom":
                all_data[couche] = await get_noeuds_telecom(db, commune, statut)
            elif couche == "liens_telecom":
                all_data[couche] = await get_liens_telecom(db, statut)
            elif couche == "logements":
                all_data[couche] = await get_logements(db, commune, statut)
            elif couche == "zones":
                all_data[couche] = await get_zones(db)
            elif couche == "noeuds_gc":
                rows = await db.fetch("SELECT id::text, nom_unique, type_noeud, etat, ST_Y(geom) AS lat, ST_X(geom) AS lng FROM noeud_gc WHERE geom IS NOT NULL")
                all_data[couche] = rows
            elif couche == "liens_gc":
                rows = await db.fetch("""SELECT lg.id::text, lg.nom_unique, lg.type_lien, lg.etat,
                       COALESCE(ST_AsGeoJSON(lg.geom), ST_AsGeoJSON(ST_MakeLine(nd.geom, na.geom)))::json AS geometry
                       FROM lien_gc lg JOIN noeud_gc nd ON nd.id=lg.id_noeud_depart JOIN noeud_gc na ON na.id=lg.id_noeud_arrivee""")
                all_data[couche] = rows
        except Exception as e:
            all_data[couche] = []

    # ─── GeoJSON ───────────────────────────────────────
    if format == "geojson":
        if len(couche_list) == 1:
            couche = couche_list[0]
            rows = all_data.get(couche, [])
            if couche in ("liens_telecom", "liens_gc", "zones"):
                gj = rows_to_geojson(rows, geom_key="geometry")
            else:
                gj = rows_to_geojson(rows)
            content = json.dumps(gj, ensure_ascii=False, indent=2)
        else:
            combined = {"type": "FeatureCollection", "features": []}
            for couche, rows in all_data.items():
                if couche in ("liens_telecom", "liens_gc", "zones"):
                    gj = rows_to_geojson(rows, geom_key="geometry")
                else:
                    gj = rows_to_geojson(rows)
                for f in gj["features"]:
                    f["properties"]["_couche"] = couche
                combined["features"].extend(gj["features"])
            content = json.dumps(combined, ensure_ascii=False, indent=2)
        return StreamingResponse(
            io.BytesIO(content.encode()),
            media_type="application/geo+json",
            headers={"Content-Disposition": f"attachment; filename=sig-ftth-export.geojson"}
        )

    # ─── CSV ────────────────────────────────────────────
    elif format == "csv":
        if len(couche_list) == 1:
            rows = all_data.get(couche_list[0], [])
            content = rows_to_csv(rows)
        else:
            parts = []
            for couche, rows in all_data.items():
                parts.append(f"# {couche}\n" + rows_to_csv(rows))
            content = "\n\n".join(parts)
        return StreamingResponse(
            io.BytesIO(content.encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=sig-ftth-export.csv"}
        )

    # ─── XLSX ────────────────────────────────────────────
    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            ORANGE = "FF6600"
            for couche, rows in all_data.items():
                if not rows: continue
                ws = wb.create_sheet(title=couche[:31])
                headers = list(dict(rows[0]).keys())
                # Header row
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=ORANGE)
                    cell.alignment = Alignment(horizontal="center")
                # Data rows
                for row_idx, row in enumerate(rows, 2):
                    for col, (k, v) in enumerate(dict(row).items(), 1):
                        val = str(v) if v is not None and not isinstance(v, (int, float, bool)) else v
                        ws.cell(row=row_idx, column=col, value=val)
                # Auto-width
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=sig-ftth-export.xlsx"}
            )
        except ImportError:
            return JSONResponse({"error": "openpyxl non disponible"}, status_code=501)

    # ─── KML ─────────────────────────────────────────────
    elif format == "kml":
        kml_parts = ['<?xml version="1.0" encoding="UTF-8"?>\n<kml xmlns="http://www.opengis.net/kml/2.2"><Document>']
        kml_parts.append('<name>SIG FTTH Orange CI</name>')
        for couche, rows in all_data.items():
            kml_parts.append(f'<Folder><name>{couche}</name>')
            for r in rows[:500]:
                d = dict(r)
                lat = d.get('lat') or d.get('latitude')
                lng = d.get('lng') or d.get('longitude')
                if lat and lng:
                    name = d.get('nom_unique', str(d.get('id','')))
                    kml_parts.append(f'<Placemark><name>{name}</name><Point><coordinates>{lng},{lat},0</coordinates></Point></Placemark>')
            kml_parts.append('</Folder>')
        kml_parts.append('</Document></kml>')
        content = '\n'.join(kml_parts)
        return StreamingResponse(
            io.BytesIO(content.encode()),
            media_type="application/vnd.google-earth.kml+xml",
            headers={"Content-Disposition": "attachment; filename=sig-ftth-export.kml"}
        )

    # ─── Shapefile (ZIP avec GeoJSON simulé) ─────────────
    elif format == "shapefile":
        try:
            import shapefile
        except ImportError:
            # Retourner GeoJSON nommé comme shapefile
            gj_all = {"type": "FeatureCollection", "features": []}
            for couche, rows in all_data.items():
                if couche in ("liens_telecom","liens_gc","zones"):
                    gj = rows_to_geojson(rows, geom_key="geometry")
                else:
                    gj = rows_to_geojson(rows)
                for f in gj["features"]:
                    f["properties"]["_couche"] = couche
                gj_all["features"].extend(gj["features"])

            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w') as zf:
                zf.writestr("sig-ftth.geojson", json.dumps(gj_all, ensure_ascii=False, indent=2))
                zf.writestr("README.txt", "Importez sig-ftth.geojson dans QGIS via Couche > Ajouter une couche vectorielle")
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/zip",
                headers={"Content-Disposition": "attachment; filename=sig-ftth-shapefile.zip"}
            )

    return JSONResponse({"error": f"Format '{format}' non supporté"}, status_code=400)


# Compatibilité avec anciens endpoints
@router.get("/geojson")
async def export_geojson(current_user: dict = Depends(get_current_user), db=Depends(get_db)):
    rows = await get_noeuds_telecom(db)
    return rows_to_geojson(rows)

@router.get("/stats")
async def stats(current_user: dict = Depends(get_current_user), db=Depends(get_db)):
    nb_nt = await db.fetchval("SELECT COUNT(*) FROM noeud_telecom") or 0
    nb_lt = await db.fetchval("SELECT COUNT(*) FROM lien_telecom") or 0
    nb_log = await db.fetchval("SELECT COUNT(*) FROM logement") or 0
    return {"noeuds_telecom": nb_nt, "liens_telecom": nb_lt, "logements": nb_log}
